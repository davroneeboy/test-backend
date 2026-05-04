import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import nested_admin
from django import forms
from django.contrib import admin, messages
from django.db.models import Count, Q
from django.contrib.auth import get_user_model
from django.contrib.auth.admin import UserAdmin as BaseUserAdmin
from django.contrib.auth.forms import UserChangeForm, UserCreationForm
from django.forms.models import BaseInlineFormSet
from django.http import HttpResponse
from django.shortcuts import redirect, render
from django.urls import path
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _

from .constants import TEST_CURATOR_GROUP, REGION_VILOYAT
from .models import (
    AdminActionLog,
    AdminActionType,
    AnswerOption,
    AttemptResponse,
    AttemptSessionEvent,
    AttemptSessionEventType,
    AttemptStatus,
    Question,
    QuestionGroup,
    Test,
    TestAttempt,
    UserProfile,
)
from .services import sync_expired_attempt

User = get_user_model()
AUDIT_LOG_VIEWERS_GROUP = "Просмотр логов админов"


def _is_test_curator(request) -> bool:
    u = request.user
    return (
        u.is_authenticated
        and not u.is_superuser
        and u.groups.filter(name=TEST_CURATOR_GROUP).exists()
    )


def _get_client_ip(request):
    forwarded = request.META.get("HTTP_X_FORWARDED_FOR")
    if forwarded:
        return forwarded.split(",")[0].strip()
    return request.META.get("REMOTE_ADDR")


class AdminActionLoggingMixin:
    def _write_admin_log(self, request, obj, action_type, changed_fields=None):
        AdminActionLog.objects.create(
            actor=request.user if request.user.is_authenticated else None,
            action_type=action_type,
            model_name=obj._meta.label,
            object_id=str(getattr(obj, "pk", "")),
            object_repr=str(obj)[:255],
            changed_fields=changed_fields or [],
            ip_address=_get_client_ip(request),
            user_agent=request.META.get("HTTP_USER_AGENT", "")[:1000],
        )

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = AdminActionType.UPDATE if change else AdminActionType.CREATE
        self._write_admin_log(request, obj, action, getattr(form, "changed_data", []))

    def delete_model(self, request, obj):
        self._write_admin_log(request, obj, AdminActionType.DELETE, [])
        super().delete_model(request, obj)

    def delete_queryset(self, request, queryset):
        for obj in queryset:
            self._write_admin_log(request, obj, AdminActionType.DELETE, [])
        super().delete_queryset(request, queryset)


class UserProfileForm(forms.ModelForm):
    class Meta:
        model = UserProfile
        fields = ("region_type", "viloyat")

    def clean(self):
        cleaned_data = super().clean()
        region_type = cleaned_data.get("region_type")
        viloyat = cleaned_data.get("viloyat")
        if region_type == REGION_VILOYAT and not viloyat:
            self.add_error("viloyat", _("Viloyatni tanlang."))
        elif region_type != REGION_VILOYAT:
            cleaned_data["viloyat"] = ""
        return cleaned_data


class UserProfileInline(admin.StackedInline):
    model = UserProfile
    form = UserProfileForm
    can_delete = False
    min_num = 1
    max_num = 1
    extra = 0
    verbose_name = _("Hudud")
    verbose_name_plural = _("Hudud")

    class Media:
        js = ("testing/admin_region.js",)


class RequiredDepartmentUserCreationForm(UserCreationForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["username"].label = _("Login")


class RequiredDepartmentUserChangeForm(UserChangeForm):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.fields["username"].label = _("Login")


class UserAdmin(AdminActionLoggingMixin, BaseUserAdmin):
    form = RequiredDepartmentUserChangeForm
    add_form = RequiredDepartmentUserCreationForm
    inlines = (UserProfileInline,)

    class Media:
        js = ("testing/admin_region.js",)

    add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "last_name",
                    "first_name",
                    "username",
                    "password1",
                    "password2",
                    "groups",
                ),
            },
        ),
    )

    _curator_add_fieldsets = (
        (
            None,
            {
                "classes": ("wide",),
                "fields": (
                    "last_name",
                    "first_name",
                    "username",
                    "password1",
                    "password2",
                    "email",
                    "groups",
                ),
            },
        ),
    )

    _curator_change_fieldsets = (
        (None, {"fields": ("username",)}),
        (_("Parol"), {"fields": ("password",)}),
        (_("Shaxsiy ma'lumotlar"), {"fields": ("first_name", "last_name", "email")}),
        (_("Bo'limlar"), {"fields": ("groups",)}),
    )

    def get_fieldsets(self, request, obj=None):
        if _is_test_curator(request):
            if obj is None:
                return self._curator_add_fieldsets
            return self._curator_change_fieldsets
        return super().get_fieldsets(request, obj)

    def save_model(self, request, obj, form, change):
        if _is_test_curator(request) and not change:
            obj.is_staff = False
            obj.is_superuser = False
        super().save_model(request, obj, form, change)


if admin.site.is_registered(User):
    admin.site.unregister(User)
admin.site.register(User, UserAdmin)


class AnswerOptionInlineFormSet(BaseInlineFormSet):
    def clean(self):
        super().clean()
        if any(self.errors):
            return
        valid_forms = [
            f for f in self.forms
            if hasattr(f, "cleaned_data")
            and not f.cleaned_data.get("DELETE", False)
            and f.cleaned_data.get("text", "").strip()
        ]
        correct_count = sum(1 for f in valid_forms if f.cleaned_data.get("is_correct"))
        if valid_forms and correct_count < 1:
            valid_forms[0].add_error(
                "is_correct",
                _("Kamida bitta to'g'ri javob belgilanishi kerak."),
            )


class AnswerOptionNestedInline(nested_admin.NestedTabularInline):
    model = AnswerOption
    extra = 4
    min_num = 2
    max_num = 10
    fields = ("text", "is_correct")
    formset = AnswerOptionInlineFormSet


class QuestionNestedInline(nested_admin.NestedStackedInline):
    model = Question
    fk_name = "group"
    extra = 0
    fields = ("order", "text", "points")
    inlines = (AnswerOptionNestedInline,)


class QuestionGroupNestedInline(nested_admin.NestedStackedInline):
    model = QuestionGroup
    extra = 0
    fields = ("department", "order", "questions_to_show")
    inlines = (QuestionNestedInline,)


class _TestImportForm(forms.Form):
    title = forms.CharField(label="Название теста", max_length=255)
    description = forms.CharField(
        label="Описание", required=False, widget=forms.Textarea(attrs={"rows": 3})
    )
    time_limit_seconds = forms.IntegerField(
        label="Лимит времени (сек)", required=False, min_value=1
    )
    questions_to_show = forms.IntegerField(
        label="Показывать вопросов", required=False, min_value=1
    )
    excel_file = forms.FileField(label="Excel-файл (.xlsx)")


@admin.register(Test)
class TestAdmin(AdminActionLoggingMixin, nested_admin.NestedModelAdmin):
    change_list_template = "admin/testing/test/change_list.html"

    list_display = (
        "title",
        "is_active",
        "conduct_schedule_display",
        "time_limit_seconds",
        "updated_at",
    )
    list_filter = ("is_active",)
    search_fields = ("title", "description")
    filter_horizontal = ("allowed_groups",)
    inlines = (QuestionGroupNestedInline,)
    readonly_fields = ("created_at", "updated_at")
    fieldsets = (
        (None, {"fields": ("title", "description", "is_active")}),
        (
            _("O'tkazish davri"),
            {
                "fields": ("conduct_starts_at", "conduct_ends_at"),
                "description": _(
                    "Bu intervaldan tashqarida test topshiruvchilarga mavjud emas. Tugagandan so'ng saqlashda «Faol» olib tashlanadi."
                ),
            },
        ),
        (_("Parametrlar"), {"fields": ("time_limit_seconds", "questions_to_show", "allowed_groups")}),
        (_("Texnik ma'lumotlar"), {"fields": ("created_at", "updated_at")}),
    )

    def get_urls(self):
        custom = [
            path(
                "import/",
                self.admin_site.admin_view(self.import_view),
                name="testing_test_import",
            ),
            path(
                "proctoring/",
                self.admin_site.admin_view(self.proctoring_view),
                name="testing_test_proctoring",
            ),
        ]
        return custom + super().get_urls()

    def proctoring_view(self, request):
        attempts = (
            TestAttempt.objects.filter(status=AttemptStatus.IN_PROGRESS)
            .select_related("user", "test")
            .order_by("started_at")
        )
        context = {
            **self.admin_site.each_context(request),
            "title": "Прокторинг",
            "attempts": attempts,
            "opts": self.model._meta,
        }
        return render(request, "admin/testing/proctoring.html", context)

    def import_view(self, request):
        from .excel_import import ParseError, import_test, parse_excel
        from .models import AdminActionType

        if request.method == "POST":
            form = _TestImportForm(request.POST, request.FILES)
            if form.is_valid():
                try:
                    rows = parse_excel(request.FILES["excel_file"])
                    test = import_test(
                        title=form.cleaned_data["title"],
                        description=form.cleaned_data.get("description") or "",
                        time_limit_seconds=form.cleaned_data.get("time_limit_seconds"),
                        questions_to_show=form.cleaned_data.get("questions_to_show"),
                        rows=rows,
                    )
                    self._write_admin_log(request, test, AdminActionType.CREATE)
                    messages.success(
                        request,
                        f"Тест «{test.title}» импортирован: {len(rows)} вопросов.",
                    )
                    return redirect(f"../{test.pk}/change/")
                except ParseError as exc:
                    messages.error(request, str(exc))
        else:
            form = _TestImportForm()

        context = {
            **self.admin_site.each_context(request),
            "title": "Импорт теста из Excel",
            "form": form,
            "opts": self.model._meta,
        }
        return render(request, "admin/testing/test/import.html", context)

    @admin.display(description=_("O'tkazish davri"))
    def conduct_schedule_display(self, obj: Test):
        from django.utils.formats import date_format

        s, e = obj.conduct_starts_at, obj.conduct_ends_at
        if not s and not e:
            return _("muddatsiz")

        def fmt(dt):
            return date_format(dt, "SHORT_DATETIME_FORMAT") if dt else _("—")

        return f"{fmt(s)} — {fmt(e)}"


class AttemptResponseInline(admin.TabularInline):
    model = AttemptResponse
    extra = 0
    readonly_fields = ("is_correct", "answered_at")

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.select_related("question", "selected_option")

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "question":
            obj_id = request.resolver_match.kwargs.get("object_id")
            if obj_id:
                try:
                    att = TestAttempt.objects.get(pk=obj_id)
                    kwargs["queryset"] = Question.objects.filter(test=att.test)
                except TestAttempt.DoesNotExist:
                    pass
        if db_field.name == "selected_option":
            obj_id = request.resolver_match.kwargs.get("object_id")
            if obj_id:
                try:
                    att = TestAttempt.objects.get(pk=obj_id)
                    kwargs["queryset"] = AnswerOption.objects.filter(
                        question__test=att.test
                    )
                except TestAttempt.DoesNotExist:
                    pass
        return super().formfield_for_foreignkey(db_field, request, **kwargs)


class AttemptSessionEventInline(admin.TabularInline):
    model = AttemptSessionEvent
    extra = 0
    can_delete = False
    max_num = 0
    fields = (
        "event_type",
        "created_at",
        "client_timestamp",
        "leave_count",
        "duration_away_ms",
        "ip_address",
    )
    readonly_fields = fields
    ordering = ("created_at",)

    def has_add_permission(self, request, obj=None):
        return False


def _fmt_duration(seconds):
    if seconds is None:
        return "—"
    m, s = divmod(int(seconds), 60)
    h, m = divmod(m, 60)
    if h:
        return f"{h}:{m:02d}:{s:02d}"
    return f"{m}:{s:02d}"


@admin.action(description=_("Tanlangan urinishlarni Excelga eksport qilish"))
def export_attempts_xlsx(modeladmin, request, queryset):
    qs = (
        queryset
        .select_related("user", "user__profile", "test")
        .prefetch_related("user__groups")
        .annotate(
            _answered=Count("responses", distinct=True),
            _total_q=Count("test__questions", distinct=True),
        )
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Urinishlar"

    _termination_labels = {
        "tab_switch": "Boshqa vkladkaga o'tganligi sababli",
        "window_blur": "Boshqa ilovaga o'tganligi sababli",
    }

    headers = [
        "Hudud turi",
        "Viloyat",
        "F.I.Sh.",
        "Login",
        "Bo'lim",
        "Test nomi",
        "Jami savollar",
        "Javob berilgan",
        "Natija (%)",
        "Yakunlanish sababi",
        "Boshlanish",
        "Tugash",
    ]

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2563EB")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for col_idx, title in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 32

    fmt_dt = "%d.%m.%Y %H:%M"

    for obj in qs:
        profile = getattr(obj.user, "profile", None)
        region_type = profile.get_region_type_display() if profile else "—"
        viloyat = (profile.get_viloyat_display() if profile and profile.viloyat else "—")
        department = obj.user.groups.first()
        pct = (
            round(float(obj.score_earned / obj.score_max * 100), 1)
            if obj.score_max
            else "—"
        )
        termination = (
            _termination_labels.get(obj.termination_reason, obj.termination_reason)
            if obj.termination_reason
            else "—"
        )
        ws.append([
            region_type,
            viloyat,
            obj.user.get_full_name() or obj.user.username,
            obj.user.username,
            department.name if department else "—",
            str(obj.test),
            obj._total_q,
            obj._answered,
            pct,
            termination,
            obj.started_at.strftime(fmt_dt) if obj.started_at else "—",
            obj.finished_at.strftime(fmt_dt) if obj.finished_at else "—",
        ])

    col_widths = [18, 24, 28, 18, 20, 32, 14, 14, 12, 32, 18, 18]
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="urinishlar.xlsx"'
    wb.save(response)
    return response


@admin.register(TestAttempt)
class TestAttemptAdmin(AdminActionLoggingMixin, admin.ModelAdmin):
    list_display = (
        "id",
        "fio_display",
        "viloyat_display",
        "test",
        "status",
        "termination_reason_display",
        "tab_hidden_events_display",
        "window_blur_events_display",
        "started_at",
        "finished_at",
        "duration_display",
        "score_display",
        "answered_count",
    )
    list_filter = (
        "status",
        "test",
        "user__groups",
        "started_at",
    )
    search_fields = (
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__email",
        "test__title",
    )
    date_hierarchy = "started_at"
    readonly_fields = (
        "started_at",
        "deadline_at",
        "finished_at",
        "termination_reason",
        "duration_display",
        "score_earned",
        "score_max",
    )
    autocomplete_fields = ("user", "test")
    inlines = (AttemptResponseInline, AttemptSessionEventInline)
    actions = ("action_sync_timeout", export_attempts_xlsx)

    def get_inline_instances(self, request, obj=None):
        inlines = super().get_inline_instances(request, obj)
        if request.user.is_superuser:
            return inlines
        if not request.user.has_perm("testing.change_attemptresponse"):
            inlines = [
                i
                for i in inlines
                if not isinstance(i, AttemptResponseInline)
            ]
        if not request.user.has_perm("testing.change_attemptsessionevent"):
            inlines = [
                i
                for i in inlines
                if not isinstance(i, AttemptSessionEventInline)
            ]
        return inlines

    @admin.display(description=_("F.I.Sh."), ordering="user__last_name")
    def fio_display(self, obj: TestAttempt):
        full = obj.user.get_full_name()
        return full or obj.user.username

    @admin.display(description=_("Viloyat"))
    def viloyat_display(self, obj: TestAttempt):
        profile = getattr(obj.user, "profile", None)
        if not profile:
            return "—"
        if profile.viloyat:
            return profile.get_viloyat_display()
        return profile.get_region_type_display()

    @admin.display(description=_("Tugatilish sababi"))
    def termination_reason_display(self, obj: TestAttempt):
        labels = {
            "tab_switch": "Boshqa vkladkaga o'tdi",
            "window_blur": "Boshqa ilovaga o'tdi",
        }
        reason = obj.termination_reason
        if not reason:
            return format_html('<span style="color:#9ca3af">—</span>')
        return format_html(
            '<span style="color:#dc2626;font-weight:600">⛔ {}</span>',
            labels.get(reason, reason),
        )

    @admin.display(description=_("Davomiyligi"))
    def duration_display(self, obj: TestAttempt):
        sec = obj.duration_seconds
        if sec is None:
            return _("—")
        if sec < 60:
            return _("{n} s").format(n=int(sec))
        m, s = divmod(int(sec), 60)
        return _("{m} daq {s} s").format(m=m, s=s)

    @admin.display(description=_("Balllar"))
    def score_display(self, obj: TestAttempt):
        if obj.score_max and obj.score_max > 0:
            pct = (obj.score_earned / obj.score_max) * 100
            pct_label = str(int(round(float(pct))))
            return format_html(
                "{} / {} (<span>{}%</span>)",
                obj.score_earned,
                obj.score_max,
                pct_label,
            )
        return f"{obj.score_earned} / {obj.score_max}"

    @admin.display(description=_("Javob berilgan savollar"))
    def answered_count(self, obj: TestAttempt):
        return obj.responses.count()

    @admin.display(
        description=_("Sahifa yashirildi"),
        ordering="_tab_hidden_count",
    )
    def tab_hidden_events_display(self, obj: TestAttempt):
        n = getattr(obj, "_tab_hidden_count", None)
        if n is None:
            n = obj.session_events.filter(
                event_type=AttemptSessionEventType.PAGE_HIDDEN
            ).count()
        if n == 0:
            return format_html('<span style="color:#9ca3af">—</span>')
        return format_html(
            '<span style="color:#b45309;font-weight:600" title="{}">⚠ {}</span>',
            _("«Sahifa yashirildi» (visibility) hodisalari qayd etildi"),
            n,
        )

    @admin.display(
        description=_("Oyna fokussiz"),
        ordering="_window_blur_count",
    )
    def window_blur_events_display(self, obj: TestAttempt):
        n = getattr(obj, "_window_blur_count", None)
        if n is None:
            n = obj.session_events.filter(
                event_type=AttemptSessionEventType.WINDOW_BLUR
            ).count()
        if n == 0:
            return format_html('<span style="color:#9ca3af">—</span>')
        return format_html(
            '<span style="color:#7c3aed;font-weight:600" title="{}">◆ {}</span>',
            _("«Oyna fokusni yo'qotdi» hodisalari qayd etildi"),
            n,
        )

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        qs = qs.select_related("user", "user__profile", "test").prefetch_related(
            "user__groups",
            "responses",
        )
        qs = qs.annotate(
            _tab_hidden_count=Count(
                "session_events",
                filter=Q(
                    session_events__event_type=AttemptSessionEventType.PAGE_HIDDEN
                ),
            ),
            _window_blur_count=Count(
                "session_events",
                filter=Q(
                    session_events__event_type=AttemptSessionEventType.WINDOW_BLUR
                ),
            ),
        )
        return qs

    @admin.action(
        description=_("Vaqt tugagandan so'ng holatni yangilash (taymaut)"),
    )
    def action_sync_timeout(self, request, queryset):
        for att in queryset:
            sync_expired_attempt(att)


@admin.register(AttemptSessionEvent)
class AttemptSessionEventAdmin(admin.ModelAdmin):
    list_display = (
        "id",
        "attempt",
        "user",
        "event_type",
        "created_at",
        "client_timestamp",
        "leave_count",
        "duration_away_ms",
        "ip_address",
    )
    list_filter = ("event_type", "created_at")
    search_fields = ("user__username", "user__email")
    date_hierarchy = "created_at"
    ordering = ("-created_at",)
    readonly_fields = (
        "attempt",
        "user",
        "event_type",
        "created_at",
        "client_timestamp",
        "duration_away_ms",
        "leave_count",
        "meta",
        "ip_address",
        "user_agent",
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return request.user.is_superuser


@admin.register(AttemptResponse)
class AttemptResponseAdmin(AdminActionLoggingMixin, admin.ModelAdmin):
    list_display = ("attempt", "question", "is_correct", "answered_at")
    list_filter = ("is_correct", "attempt__test")
    search_fields = ("attempt__user__username", "question__text")
    readonly_fields = ("answered_at",)



@admin.register(AdminActionLog)
class AdminActionLogAdmin(admin.ModelAdmin):
    list_display = (
        "created_at",
        "actor",
        "action_type",
        "model_name",
        "object_id",
        "object_repr",
    )
    list_filter = ("action_type", "model_name", "created_at")
    search_fields = ("actor__username", "model_name", "object_id", "object_repr")
    date_hierarchy = "created_at"
    readonly_fields = (
        "actor",
        "action_type",
        "model_name",
        "object_id",
        "object_repr",
        "changed_fields",
        "ip_address",
        "user_agent",
        "created_at",
    )

    def has_add_permission(self, request):
        return False

    def has_change_permission(self, request, obj=None):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

    def _can_view(self, request):
        user = request.user
        if not (user and user.is_active and user.is_staff):
            return False
        if user.is_superuser:
            return True
        return bool(
            user.groups.filter(name=AUDIT_LOG_VIEWERS_GROUP).exists()
            or user.has_perm("testing.view_adminactionlog")
        )

    def has_module_permission(self, request):
        return self._can_view(request)

    def has_view_permission(self, request, obj=None):
        return self._can_view(request)
